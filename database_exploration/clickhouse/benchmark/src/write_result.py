import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Color, NamedStyle, PatternFill, Side, colors


class Excel_Writer:
    def __init__(self):
        self.filename = "./output/benchmark_results.xlsx"
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "WRITE"
        self.sheet["A1"] = "row_count"
        self.sheet["B1"] = "batch_length"
        self.sheet["C1"] = "insert_time"
        self.sheet["D1"] = "speed"
        self.sheet["E1"] = "shard 1"
        self.sheet["F1"] = "shard 2"
        self.sheet["R1"] = "батчи по 10.000"
        self.sheet["U1"] = "батчи по 100.000"
        self.sheet.column_dimensions["A"].adjust_column_width = True
        self.sheet.column_dimensions["B"].adjust_column_width = True
        self.sheet.column_dimensions["C"].adjust_column_width = True
        self.sheet.column_dimensions["D"].adjust_column_width = True
        self.sheet.column_dimensions["E"].adjust_column_width = True
        self.sheet.column_dimensions["F"].adjust_column_width = True
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        alignment = Alignment(horizontal="center", vertical="center")
        style = NamedStyle(name="header", border=border, alignment=alignment)
        self.sheet["A1"].style = style
        self.sheet["B1"].style = style
        self.sheet["C1"].style = style
        self.sheet["D1"].style = style
        self.sheet["E1"].style = style
        self.sheet["F1"].style = style
        fill = PatternFill(patternType="solid", fgColor=colors.BLUE)
        self.sheet["Q1"].fill = fill
        fill = PatternFill(patternType="solid", fgColor=Color(rgb="00FF0000"))
        self.sheet["T1"].fill = fill

        self.sheet_1 = self.workbook.create_sheet(title="REED")
        self.sheet_1["A1"] = "rows_count"
        self.sheet_1["B1"] = "speed"
        self.sheet_1["C1"] = "stress"
        self.sheet_1["D1"] = "all data DB cnt"
        self.sheet_1.column_dimensions["A"].adjust_column_width = True
        self.sheet_1.column_dimensions["B"].adjust_column_width = True
        self.sheet_1.column_dimensions["C"].adjust_column_width = True
        self.sheet_1.column_dimensions["D"].adjust_column_width = True
        self.sheet_1["A1"].style = style
        self.sheet_1["B1"].style = style
        self.sheet_1["C1"].style = style
        self.sheet_1["D1"].style = style

        self.workbook.save(self.filename)

    def write_load_result(self, row_count, batch_length, insert_time, speed, shard_1, shard_2):
        self.sheet.append([row_count, batch_length, insert_time, speed, shard_1, shard_2])
        self.update_charts()
        self.workbook.save(self.filename)

    def write_reed_result(self, count, speed, stress, all_data):
        self.sheet_1.append([count, speed, stress, all_data])
        self.update_charts_sheet_2()
        self.workbook.save(self.filename)

    def update_charts_sheet_2(self):
        chart = BarChart()
        data = Reference(
            self.sheet_1, min_col=2, max_col=2, min_row=1, max_row=self.sheet_1.max_row
        )
        chart.add_data(data, titles_from_data=True)
        cats = Reference(
            self.sheet_1, min_col=1, min_row=2, max_col=1, max_row=self.sheet_1.max_row
        )
        chart.set_categories(cats)
        chart.title = "Скорость чтения"
        chart.x_axis.title = "кол-во записей"
        chart.y_axis.title = "скорость записи/сек"
        chart.grouping = "standard"
        self.sheet_1.add_chart(chart, "H1")

    def update_charts(self):
        chart = BarChart()
        data = Reference(self.sheet, min_col=4, max_col=4, min_row=1, max_row=7)
        chart.add_data(data, titles_from_data=True)
        cats = Reference(self.sheet, min_col=1, min_row=2, max_col=1, max_row=7)
        chart.set_categories(cats)
        chart.title = "Скорость при батче 10.000"
        chart.x_axis.title = "кол-во записей"
        chart.y_axis.title = "скорость записи/сек"
        chart.grouping = "standard"
        self.sheet.add_chart(chart, "H1")

        chart1 = BarChart()
        data = Reference(self.sheet, min_col=4, max_col=4, min_row=7, max_row=13)
        chart1.add_data(data, titles_from_data=True)
        cats = Reference(self.sheet, min_col=1, max_col=1, min_row=8, max_row=13)
        chart1.set_categories(cats)
        chart1.title = "Скорость при батче 100.000"
        chart1.x_axis.title = "кол-во записей"
        chart1.y_axis.title = "скорость записи/сек"
        chart1.grouping = "standard"
        self.sheet.add_chart(chart1, "H16")

        chart2 = LineChart()
        chart2.title = "сравнение скоростей при разных батчах"
        chart2.x_axis.title = "кол-во записей"
        chart2.y_axis.title = "скорость записи/сек"
        x_data = Reference(self.sheet, min_col=1, min_row=2, max_row=13)
        chart2.set_categories(x_data)
        y1_data = Reference(self.sheet, min_col=4, min_row=1, max_row=7)
        y2_data = Reference(self.sheet, min_col=4, min_row=7, max_row=13)
        chart2.add_data(y1_data, titles_from_data=True)
        chart2.add_data(y2_data, titles_from_data=True)
        self.sheet.add_chart(chart2, "Q2")
